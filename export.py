# =============================================================
#  export.py  —  Export SQLite → Excel with colour coding
#  Run: python export.py
# =============================================================
import sqlite3
from pathlib import Path
from datetime import datetime
from config import DB_PATH, EXPORT_DIR


def export_to_excel(out: str = None) -> str:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  Run: pip install openpyxl"); return ""

    Path(EXPORT_DIR).mkdir(parents=True, exist_ok=True)
    if not out:
        out = str(Path(EXPORT_DIR) /
                  f"zoho_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("""
        SELECT sno, file_name, image1_path, image2_path,
               file_order_id, file_star,
               remarks_file_order_id, remarks_file_star,
               flag, ocr_engine_order, ocr_engine_star, processed_at
        FROM zoho_records ORDER BY sno
    """).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Zoho Records"

    if not rows:
        wb.save(out); return out

    headers = list(rows[0].keys())
    hfill   = PatternFill("solid", fgColor="1F4E79")
    hfont   = Font(color="FFFFFF", bold=True)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h.upper().replace("_"," "))
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal="center")

    COLOURS = {
        "OK":           "C6EFCE",   # green
        "PENDING":      "FFEB9C",   # yellow
        "ERROR":        "FFC7CE",   # red
        "NO_ORDER_ID":  "FCE4D6",   # orange
        "NO_STAR":      "FCE4D6",
        "LOW_CONF":     "DDEBF7",   # light blue
        "MISSING":      "D9D9D9",   # grey
    }
    fcol = headers.index("flag") + 1 if "flag" in headers else None

    for ri, row in enumerate(rows, 2):
        d = dict(row)
        for ci, h in enumerate(headers, 1):
            ws.cell(row=ri, column=ci, value=d.get(h))
        if fcol:
            fv  = d.get("flag") or ""
            clr = next((v for k, v in COLOURS.items() if k in fv), "FFFFFF")
            ws.cell(row=ri, column=fcol).fill = PatternFill("solid", fgColor=clr)

    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 60)

    wb.save(out)
    print(f"  Exported {len(rows):,} rows → {out}")
    return out


if __name__ == "__main__":
    export_to_excel()
